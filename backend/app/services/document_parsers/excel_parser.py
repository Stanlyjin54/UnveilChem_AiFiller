#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文档解析器 - 使用pandas和openpyxl
"""

import os
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
import json

try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("Excel解析依赖缺失，请安装pandas, openpyxl")

from . import BaseDocumentParser

class ExcelParser(BaseDocumentParser):
    """Excel文档解析器"""
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
        self.parser_name = "EXCEL_PARSER_V1"
    
    def can_parse(self, file_path: Path) -> bool:
        """检查是否是Excel文件"""
        return file_path.suffix.lower() in ['.xlsx', '.xls']
    
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """解析Excel文件"""
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "error": "Excel解析功能不可用，请安装pandas, openpyxl依赖",
                "file_path": str(file_path),
                "parser_used": self.parser_name,
                "metadata": self.get_metadata(file_path)
            }
        
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        result = {
            "success": False,
            "parser_used": self.parser_name,
            "file_path": str(file_path),
            "metadata": self.get_metadata(file_path),
            "sheets": [],
            "tables": [],
            "parameters": {},
            "errors": []
        }
        
        try:
            # 使用openpyxl获取工作表信息
            workbook = load_workbook(str(file_path), read_only=True, data_only=True)
            
            result["metadata"]["sheet_names"] = workbook.sheetnames
            result["metadata"]["sheets_count"] = len(workbook.sheetnames)
            
            # 逐个处理工作表
            for sheet_name in workbook.sheetnames:
                sheet_info = self._parse_sheet(workbook, sheet_name)
                if sheet_info:
                    result["sheets"].append(sheet_info)
            
            workbook.close()
            
            # 合并所有文本内容用于参数提取
            all_text = ""
            for sheet in result["sheets"]:
                all_text += sheet.get("text_summary", "") + "\n"
            
            # 提取参数
            if all_text.strip():
                result["parameters"] = self._extract_parameters(all_text)
            
            result["success"] = True
            result["metadata"]["total_tables"] = len(result["tables"])
            result["metadata"]["total_cells"] = sum(sheet.get("cell_count", 0) for sheet in result["sheets"])
            
        except Exception as e:
            result["errors"].append(f"Excel解析失败: {str(e)}")
            logging.error(f"Excel解析错误: {e}")
        
        return result
    
    def _parse_sheet(self, workbook, sheet_name: str) -> Dict[str, Any]:
        """解析单个工作表"""
        try:
            worksheet = workbook[sheet_name]
            
            # 获取工作表基本信息
            sheet_info = {
                "name": sheet_name,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "cell_count": 0,
                "data": [],
                "headers": [],
                "parameters": {},
                "has_data": False
            }
            
            # 读取数据
            all_rows = []
            for row_idx in range(1, min(worksheet.max_row + 1, 1000)):  # 限制最多1000行
                row_data = []
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = self._get_cell_value(cell)
                    row_data.append(cell_value)
                    if cell_value is not None and str(cell_value).strip():
                        sheet_info["cell_count"] += 1
                
                # 过滤空行
                if any(cell is not None and str(cell).strip() for cell in row_data):
                    all_rows.append(row_data)
            
            sheet_info["data"] = all_rows
            
            # 提取表头（第一行）
            if all_rows:
                sheet_info["headers"] = [str(cell) if cell is not None else "" for cell in all_rows[0]]
                sheet_info["has_data"] = True
                
                # 参数提取
                text_content = self._sheet_to_text(all_rows)
                if text_content.strip():
                    sheet_info["parameters"] = self._extract_parameters(text_content)
                    
                    # 检查是否包含工艺参数数据
                    if sheet_info["parameters"]:
                        sheet_info["contains_parameters"] = True
            
            return sheet_info
            
        except Exception as e:
            logging.error(f"工作表 {sheet_name} 解析失败: {e}")
            return None
    
    def _get_cell_value(self, cell) -> Any:
        """安全获取单元格值"""
        try:
            if cell.value is None:
                return None
            
            # 处理日期
            from datetime import datetime
            if isinstance(cell.value, datetime):
                return cell.value.strftime("%Y-%m-%d %H:%M:%S")
            
            # 处理数字
            if isinstance(cell.value, (int, float)):
                return cell.value
            
            # 转换为字符串并清理
            return str(cell.value).strip()
            
        except Exception:
            return str(cell.value) if cell.value is not None else None
    
    def _sheet_to_text(self, rows: List[List[Any]]) -> str:
        """将工作表数据转换为文本"""
        text_lines = []
        for row in rows:
            line_parts = []
            for cell in row:
                if cell is not None and str(cell).strip():
                    line_parts.append(str(cell))
            if line_parts:
                text_lines.append(" ".join(line_parts))
        return "\n".join(text_lines)
    
    def _extract_parameters(self, text: str) -> Dict[str, Any]:
        """提取工艺参数"""
        import re
        parameters = {}
        
        # 温度参数
        temp_patterns = [
            r'(?:温度|Temperature|TEMP)[:\s]*([\d.]+)\s*[°℃]?C?',
            r'([\d.]+)\s*[°℃](?:\s*温度|\s*Temperature)?',
            r'T\s*[=]\s*([\d.]+)\s*[°℃]?',
            r'Temp(?:\.|erature)?[:\s]*([\d.]+)\s*[°℃]?C?'
        ]
        
        # 压力参数
        pressure_patterns = [
            r'(?:压力|Pressure|PRESS)[:\s]*([\d.]+)\s*(?:MPa|kPa|Pa|Bar)?',
            r'([\d.]+)\s*(?:MPa|kPa|Pa|Bar)(?:\s*压力|\s*Pressure)?',
            r'P\s*[=]\s*([\d.]+)\s*(?:MPa|kPa|Pa)?'
        ]
        
        # 流量参数
        flow_patterns = [
            r'(?:流量|Flow|FLOW)[:\s]*([\d.]+)\s*(?:m³/h|L/min|kg/h|L/h)?',
            r'([\d.]+)\s*(?:m³/h|L/min|kg/h|L/h)(?:\s*流量|\s*Flow)?',
            r'Flow\s*rate[:\s]*([\d.]+)\s*(?:m³/h|L/min|kg/h)?'
        ]
        
        # 浓度参数
        concentration_patterns = [
            r'(?:浓度|Concentration|CONC)[:\s]*([\d.]+)\s*(?:%|ppm|mg/L|g/L)?',
            r'([\d.]+)\s*(?:%|ppm|mg/L|g/L)(?:\s*浓度|\s*Concentration)?'
        ]
        
        # 提取各类参数
        for pattern_list, param_type in [
            (temp_patterns, "temperature"),
            (pressure_patterns, "pressure"),
            (flow_patterns, "flow"),
            (concentration_patterns, "concentration")
        ]:
            values = []
            for pattern in pattern_list:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    try:
                        value = float(match.group(1))
                        if self._validate_parameter_value(param_type, value):
                            values.append({
                                "value": value,
                                "unit": self._extract_unit(match.group(0), param_type),
                                "context": match.group(0).strip(),
                                "position": match.span()
                            })
                    except (ValueError, IndexError):
                        continue
            
            if values:
                parameters[param_type] = values
        
        return parameters
    
    def _validate_parameter_value(self, param_type: str, value: float) -> bool:
        """验证参数值是否合理"""
        validation_rules = {
            "temperature": (-100, 1000),  # -100°C 到 1000°C
            "pressure": (0, 50),          # 0 到 50 MPa
            "flow": (0, 10000),           # 0 到 10000 m³/h
            "concentration": (0, 100)     # 0% 到 100%
        }
        
        if param_type in validation_rules:
            min_val, max_val = validation_rules[param_type]
            return min_val <= value <= max_val
        
        return True
    
    def _extract_unit(self, text: str, param_type: str) -> str:
        """提取参数单位"""
        import re
        
        unit_patterns = {
            "temperature": [r'°C', r'℃', r'C(?![a-zA-Z])'],
            "pressure": [r'MPa', r'kPa', r'Pa', r'Bar'],
            "flow": [r'm³/h', r'L/min', r'kg/h', r'L/h'],
            "concentration": [r'%|ppm|mg/L|g/L']
        }
        
        if param_type in unit_patterns:
            for pattern in unit_patterns[param_type]:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    return match.group(0)
        
        # 默认单位
        default_units = {
            "temperature": "°C",
            "pressure": "MPa", 
            "flow": "m³/h",
            "concentration": "%"
        }
        
        return default_units.get(param_type, "")